'''

Copy this file to the directory where the csv files are present and run it

It will convert all csv files into xlsx and moves the csv files into the ....csv... folder which will be created in run time

'''



import os
import glob
import csv
import openpyxl
from openpyxl import load_workbook
import shutil

for csvfile in glob.glob(os.path.join('.', '*.csv')):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('Sheet1')
    with open(csvfile, 'rb') as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader, start=1):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c).value = val
    newfile = csvfile.replace('csv','xlsx')
    print newfile
    del wb['Sheet']
    print wb.sheetnames
    wb.save(newfile)
for csvfile in glob.glob(os.path.join('.', '*.csv')):
    if not os.path.exists('./CSV'):
        os.makedirs('./CSV')
    shutil.move(csvfile, './CSV')    
#     os.remove(csvfile)
        
    
