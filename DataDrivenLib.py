import shutil;
import os;
import xlrd
from xlrd.formula import sheetrange
import itertools
from itertools import permutations, count
from sets import Set
from xlutils.copy import copy
import xlwt
import xlsxwriter
from openpyxl import Workbook
import openpyxl
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import json
def Update_Multiple_Combinations_of_Data(InputDataSheetPath,OutputDatasheet,InputsheetName='Sheet1',OutputSheetName='Sheet1'):
    workbook = xlrd.open_workbook(InputDataSheetPath)
    sheet = workbook.sheet_by_name(InputsheetName)

    list = []
    cart = []
    perm = [] 

    for col in range(0, sheet.ncols):
        list=[]
        for row in range (1,sheet.nrows):
            val = sheet.cell(row,col).value
            if isinstance(val,unicode) == True:
                val=str(val)   
            list.append(val)
        cart.append(list)

#     book = xlwt.Workbook(encoding="utf-8")
    book = xlsxwriter.Workbook(OutputDatasheet)
    sheet1 = book.add_worksheet(OutputSheetName)
#     sheet1 = book.add_sheet(OutputSheetName, cell_overwrite_ok=True)
    for comb in itertools.product(*cart):
        perm.append(comb)

    result_tuple = perm[0]
    len_tuple = len(result_tuple)
    len_list = len(perm)
    print len_tuple
    print len_list
    
    print perm
    Columnheader = _getHeaderColumnfunc(InputDataSheetPath, InputsheetName)
    Columnheader.append("Status")
#     if rowid not in Columnheader:
    Columnheader.insert(0, "rowid")
#     count = "0"
    print "ColumnHeaderof the Excel",Columnheader
#      for copying the Header
    for i in range(0,1):
        for j in range(0,len(Columnheader)):
            sheet1.write(i, j, Columnheader[int(j)])
#     for copying the Rows                   
    for i in range(1,len_list+1):
        for j in range(1,len_tuple+1):
#             count=int(count)+1
            sheet1.write(i, 0, str(i))
            sheet1.write(i, j, perm[int(i-1)][int(j-1)])
        
    book.close()
    return OutputDatasheet
    
    
def _getHeaderColumnfunc(InputDataSheetPath,InputsheetName='Sheet1'):
    columnheader =[]
    workbook = xlrd.open_workbook(InputDataSheetPath)
    sheet = workbook.sheet_by_name(InputsheetName)
    rows = sheet.nrows
    print "Number of rows",rows
    columns = sheet.ncols
    for i in range(1):
        for j in range(columns):
            columnheader.append(sheet.cell(i,j).value)
    return columnheader    
    workbook.release_resources()
    
def WriteIntoExcel(filepath,Row,columnvalue,columnname="Status",sheetName="Sheet1"):
#     wb = Workbook(filepath)
#     nameslist = wb.sheetnames
#     print nameslist
#     ws = wb.get_sheet_by_name(sheetName)
#     columnnum = _getRequiredHeader(filepath, columnname,sheetName)   
#     ws.append([int(Row),columnnum,columnvalue])
#     wb.save(filepath)
    Row = int(Row) + 1
    xfile = openpyxl.load_workbook(filepath)

    sheet = xfile[sheetName]
    print "sheetdata",sheet
#     sheet['A1'] = 'hello world'
    columnnum = _getRequiredHeader(filepath, columnname,sheetName)
    print "column number is",columnnum
    sheet.cell(row=int(Row), column=int(columnnum), value=str(columnvalue))
#     colVal = sheet.cell(column=columnnum, row=Row, value="{0}".format(get_column_letter(columnnum)))
#     print "column Value is",colVal
#     sheet[colVal] = str(columnvalue)
    xfile.save(filepath)
    
    
    
#     workbook1 = xlsxwriter.Workbook(filepath)
#     print "debug",workbook1
#     sheet1 = workbook1.worksheets()
#     print "sheet1",sheet1
#     sheet1 = workbook1.get_worksheet_by_name(sheetName)
#     workbook1 = xlrd.open_workbook(filepath)
#     writeworkbook1 = copy(sheet1)
#     sheet = writeworkbook1.get_sheet(0)
#     columnnum = _getRequiredHeader(filepath, columnname,sheetName)   
#     sheet1.write(int(Row),columnnum,columnvalue)
#     writeworkbook1.save(filepath)    
    
def _getRequiredHeader(filepath,columnname,sheetName):
    headerExists = False
    headerrow = []
    headerrow = _getHeaderColumnfunc(filepath,sheetName)
    l = len(headerrow)
    print "Len",l
    for i in range(l):
        if(headerrow[i]==columnname):
            headerExists = True
            break
        if(i == l-1):
            if(headerExists == False): raise AssertionError("Entered Header not found in the sheet : "+columnname)       
    return i+1     
    
    
